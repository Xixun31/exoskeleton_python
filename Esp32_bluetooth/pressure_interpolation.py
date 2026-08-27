import serial
import time
import numpy as np
import pandas as pd


# ============================================================
# 使用者設定
# ============================================================

# 改成你的 STM32 COM Port
PORT = "COM9"

BAUD = 115200


# ------------------------------------------------------------
# 採集模式
#
# True  = 固定收幾筆，例如500筆
# False = 固定收幾秒，例如30秒
# ------------------------------------------------------------

USE_PACKET_COUNT = True


# 固定筆數模式
TARGET_RAW_SAMPLES = 500


# 固定時間模式
RECORD_SECONDS = 30


# ------------------------------------------------------------
# 內插目標頻率
# ------------------------------------------------------------

TARGET_INTERP_HZ = 60.0


# ------------------------------------------------------------
# Excel檔名
# ------------------------------------------------------------

from datetime import datetime

OUTPUT_FILE = (
    "foot_pressure_20Hz_vs_60Hz_"
    + datetime.now().strftime("%Y%m%d_%H%M%S")
    + ".xlsx"
)

# ============================================================
# COP 感測區域幾何中心 (AutoCAD 世界座標, mm)
# ============================================================
sensor_x = np.array([
    1943.9325, 1940.0133, 1973.6715, 1972.9752, 1965.7372, 1957.7049,
    1963.0637, 1955.6792, 1995.8721, 1996.5690, 1983.1920, 1980.9491,
    1982.3130, 1970.9490, 2000.9966, 2003.7015, 2001.4963, 1988.7695
], dtype=float)

sensor_y = np.array([
    2600.1604, 2637.9524, 2444.2083, 2482.2769, 2521.7471, 2561.9473,
    2599.0265, 2642.2744, 2444.2212, 2482.2769, 2520.6611, 2559.8462,
    2599.0265, 2641.6030, 2521.0309, 2560.0963, 2597.5737, 2633.6218
], dtype=float)

# 原點尚未決定，先設 0。之後只需修改這兩個值。
ORIGIN_X = 1984.7721
ORIGIN_Y = 2415.3761

# 總壓力過低時不計算 COP（先暫定 50 g）
FORCE_THRESHOLD = 50.0

def add_cop_columns(df):
    pressure_cols = [f"P{i}" for i in range(1, 19)]
    F = df[pressure_cols].to_numpy(dtype=float)
    total = F.sum(axis=1)

    x_rel = sensor_x - ORIGIN_X
    y_rel = sensor_y - ORIGIN_Y

    cop_x = np.full(len(df), np.nan, dtype=float)
    cop_y = np.full(len(df), np.nan, dtype=float)

    valid = total >= FORCE_THRESHOLD
    if np.any(valid):
        cop_x[valid] = (F[valid] @ x_rel) / total[valid]
        cop_y[valid] = (F[valid] @ y_rel) / total[valid]

    df["COP_X_mm"] = cop_x
    df["COP_Y_mm"] = cop_y
    df["COP_valid"] = valid
    return df


# ============================================================
# 開啟 Serial
# ============================================================

ser = serial.Serial(
    PORT,
    BAUD,
    timeout=1
)


# 清除可能已經存在buffer裡的舊資料
ser.reset_input_buffer()


print("========================================")
print("Foot Pressure Data Recorder")
print("========================================")
print(f"COM Port          : {PORT}")
print(f"Baud              : {BAUD}")

if USE_PACKET_COUNT:
    print(
        f"Recording mode    : "
        f"{TARGET_RAW_SAMPLES} samples"
    )
else:
    print(
        f"Recording mode    : "
        f"{RECORD_SECONDS} seconds"
    )

print(
    f"Interpolation     : "
    f"{TARGET_INTERP_HZ:.0f} Hz"
)

print()
print("等待 STM32 DATA...")
print()


# ============================================================
# 儲存原始資料
# ============================================================

raw_data = []

pc_start_time = None


# ============================================================
# 接收資料
# ============================================================

try:

    while True:

        line = ser.readline().decode(
            "utf-8",
            errors="ignore"
        ).strip()


        # STM32可能還會印：
        #
        # Sensor -> ESP32 -> STM32
        # Checksum Error...
        #
        # 我們只要DATA開頭
        # ----------------------------------------------------

        if not line.startswith("DATA,"):
            continue


        parts = line.split(",")


        # 正常應該是：
        #
        # DATA
        # timestamp
        # P1~P18
        #
        # 合計20欄
        # ----------------------------------------------------

        if len(parts) != 20:

            print(
                "格式錯誤，忽略：",
                line
            )

            continue


        try:

            stm32_time_ms = int(
                parts[1]
            )

            pressures = [
                int(value)
                for value
                in parts[2:20]
            ]

        except ValueError:

            print(
                "數值解析錯誤，忽略：",
                line
            )

            continue


        # ----------------------------------------------------
        # 收到第一筆
        # ----------------------------------------------------

        if pc_start_time is None:

            pc_start_time = (
                time.perf_counter()
            )

            print(
                "收到第一筆資料，開始紀錄..."
            )


        # ----------------------------------------------------
        # 建立一列
        # ----------------------------------------------------

        row = {
            "time_ms":
                stm32_time_ms,

            "time_s":
                stm32_time_ms / 1000.0
        }


        total_pressure = 0


        for i in range(18):

            value = pressures[i]

            row[f"P{i+1}"] = value

            total_pressure += value


        row["Total"] = total_pressure


        raw_data.append(row)


        count = len(raw_data)


        # ----------------------------------------------------
        # 每20筆顯示一次
        # ----------------------------------------------------

        if count % 20 == 0:

            print(
                f"Raw samples: "
                f"{count}"
            )


        # ====================================================
        # 停止條件
        # ====================================================

        if USE_PACKET_COUNT:

            if (
                count >=
                TARGET_RAW_SAMPLES
            ):
                break


        else:

            elapsed_pc = (
                time.perf_counter()
                - pc_start_time
            )

            if (
                elapsed_pc >=
                RECORD_SECONDS
            ):
                break


except KeyboardInterrupt:

    print()
    print("使用者手動停止")


finally:

    ser.close()


# ============================================================
# 是否有收到足夠資料
# ============================================================

if len(raw_data) < 2:

    print(
        "資料太少，無法進行內插"
    )

    raise SystemExit


print()
print("========================================")
print("原始資料接收完成")
print("========================================")
print(
    f"Raw samples : "
    f"{len(raw_data)}"
)


# ============================================================
# 建立原始 DataFrame
# ============================================================

raw_df = pd.DataFrame(
    raw_data
)


# ============================================================
# 把第一筆時間改成0
#
# 原本例如：
# 1234, 1286, 1338...
#
# 改成：
# 0, 52, 104...
#
# 比較容易畫圖
# ============================================================

raw_start_ms = (
    raw_df["time_ms"].iloc[0]
)


raw_df["relative_time_ms"] = (
    raw_df["time_ms"]
    - raw_start_ms
)


raw_df["relative_time_s"] = (
    raw_df["relative_time_ms"]
    / 1000.0
)


# ============================================================
# 計算真實20Hz資料頻率
# ============================================================

raw_duration = (
    raw_df[
        "relative_time_s"
    ].iloc[-1]
)


if raw_duration > 0:

    raw_actual_hz = (
        (len(raw_df) - 1)
        /
        raw_duration
    )

else:

    raw_actual_hz = 0


# ============================================================
# 每兩個原始點之間固定插入兩個點
# 也就是 3x linear interpolation
#
# 注意：
# 原始資料實測約 19.2 Hz，因此 3 倍後約為 57.6 Hz，
# 不一定是精準 60.0 Hz。
# ============================================================

interp_rows = []

for k in range(len(raw_df) - 1):

    row0 = raw_df.iloc[k]
    row1 = raw_df.iloc[k + 1]

    t0 = row0["relative_time_s"]
    t1 = row1["relative_time_s"]

    # --------------------------------------------------------
    # 先加入原始量測點
    # --------------------------------------------------------

    new_row = {
        "relative_time_s": t0,
        "relative_time_ms": t0 * 1000.0,
        "source": "Measured"
    }

    for i in range(1, 19):
        new_row[f"P{i}"] = row0[f"P{i}"]

    interp_rows.append(new_row)

    # --------------------------------------------------------
    # 在相鄰兩個原始點之間補兩個線性內插點
    # alpha = 1/3, 2/3
    # --------------------------------------------------------

    for alpha in (1.0 / 3.0, 2.0 / 3.0):

        t = t0 + alpha * (t1 - t0)

        new_row = {
            "relative_time_s": t,
            "relative_time_ms": t * 1000.0,
            "source": "Interpolated"
        }

        for i in range(1, 19):

            f0 = row0[f"P{i}"]
            f1 = row1[f"P{i}"]

            new_row[f"P{i}"] = (
                f0 + alpha * (f1 - f0)
            )

        interp_rows.append(new_row)


# ============================================================
# 補上最後一個原始量測點
# ============================================================

last = raw_df.iloc[-1]

last_row = {
    "relative_time_s": last["relative_time_s"],
    "relative_time_ms": last["relative_time_s"] * 1000.0,
    "source": "Measured"
}

for i in range(1, 19):
    last_row[f"P{i}"] = last[f"P{i}"]

interp_rows.append(last_row)


# ============================================================
# 建立內插後 DataFrame
# ============================================================

interp_df = pd.DataFrame(interp_rows)


# ============================================================
# Total
# ============================================================

pressure_columns = [
    f"P{i}"
    for i in range(1, 19)
]

interp_df["Total"] = (
    interp_df[pressure_columns]
    .sum(axis=1)
)


# ============================================================
# 標記來源
# ============================================================

raw_df["source"] = (
    "Measured"
)

# interp_df 的 source 已在建立每一列時標記為
# Measured 或 Interpolated，因此這裡不再覆寫。


# ============================================================
# 計算原始約20Hz與3x內插後約60Hz的 COP
# ============================================================
raw_df = add_cop_columns(raw_df)
interp_df = add_cop_columns(interp_df)


# ============================================================
# COP 比較指標
# 在原始量測時間點，內插資料保留原始值，因此理論上差異接近 0
# ============================================================
interp_measured_df = interp_df[interp_df["source"] == "Measured"].reset_index(drop=True)
raw_cmp = raw_df.reset_index(drop=True)

compare_len = min(len(raw_cmp), len(interp_measured_df))
if compare_len > 0:
    dx = (interp_measured_df.loc[:compare_len-1, "COP_X_mm"].to_numpy()
          - raw_cmp.loc[:compare_len-1, "COP_X_mm"].to_numpy())
    dy = (interp_measured_df.loc[:compare_len-1, "COP_Y_mm"].to_numpy()
          - raw_cmp.loc[:compare_len-1, "COP_Y_mm"].to_numpy())
    cop_point_error = np.sqrt(dx**2 + dy**2)
    mean_cop_error_mm = np.nanmean(cop_point_error) if np.any(~np.isnan(cop_point_error)) else np.nan
    max_cop_error_mm = np.nanmax(cop_point_error) if np.any(~np.isnan(cop_point_error)) else np.nan
else:
    mean_cop_error_mm = np.nan
    max_cop_error_mm = np.nan


# ============================================================
# 3x 內插後實際頻率
# ============================================================

if (
    len(interp_df) > 1
):

    interp_duration = (
        interp_df[
            "relative_time_s"
        ].iloc[-1]
        -
        interp_df[
            "relative_time_s"
        ].iloc[0]
    )

    interp_actual_hz = (
        (len(interp_df) - 1)
        /
        interp_duration
    )

else:

    interp_duration = 0
    interp_actual_hz = 0


# ============================================================
# Summary
# ============================================================

summary_df = pd.DataFrame({

    "Item": [
        "Raw samples",
        "Raw duration (s)",
        "Raw actual frequency (Hz)",
        "Interpolated samples",
        "Interpolated duration (s)",
        "Nominal target frequency (Hz)",
        "3x interpolated actual frequency (Hz)",
        "Mean COP difference at measured timestamps (mm)",
        "Max COP difference at measured timestamps (mm)"
    ],

    "Value": [
        len(raw_df),
        raw_duration,
        raw_actual_hz,
        len(interp_df),
        interp_duration,
        TARGET_INTERP_HZ,
        interp_actual_hz,
        mean_cop_error_mm,
        max_cop_error_mm
    ]
})


# ============================================================
# 匯出Excel
# ============================================================

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    # --------------------------------------------------------
    # 原始20Hz
    # --------------------------------------------------------

    raw_df.to_excel(
        writer,
        sheet_name="Raw_20Hz",
        index=False
    )


    # --------------------------------------------------------
    # 內插60Hz
    # --------------------------------------------------------

    interp_df.to_excel(
        writer,
        sheet_name="Interp_60Hz",
        index=False
    )


    # --------------------------------------------------------
    # Summary
    # --------------------------------------------------------

    summary_df.to_excel(
        writer,
        sheet_name="Summary",
        index=False
    )

    # --------------------------------------------------------
    # Excel 圖表：COP-X / COP-Y / COP trajectory
    # --------------------------------------------------------
    from openpyxl.chart import ScatterChart, Reference, Series

    wb = writer.book
    ws_raw = wb["Raw_20Hz"]
    ws_interp = wb["Interp_60Hz"]

    # 找欄位位置
    raw_headers = {cell.value: cell.column for cell in ws_raw[1]}
    interp_headers = {cell.value: cell.column for cell in ws_interp[1]}

    def add_xy_series(chart, ws, x_col, y_col, title, max_row, marker=None, smooth=False):
        xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=max_row)
        yvalues = Reference(ws, min_col=y_col, min_row=2, max_row=max_row)
        series = Series(yvalues, xvalues, title=title)
        if marker is not None:
            series.marker.symbol = marker
            series.marker.size = 5
        if smooth:
            series.smooth = True
        chart.series.append(series)

    # COP-X vs time
    chart_x = ScatterChart()
    chart_x.title = "COP-X: Raw ~20Hz vs 3x Interpolated ~60Hz"
    chart_x.x_axis.title = "Time (s)"
    chart_x.y_axis.title = "COP-X (mm)"
    chart_x.height = 10
    chart_x.width = 18
    add_xy_series(chart_x, ws_raw, raw_headers["relative_time_s"], raw_headers["COP_X_mm"],
                  "Raw ~20Hz", ws_raw.max_row, marker="circle")
    add_xy_series(chart_x, ws_interp, interp_headers["relative_time_s"], interp_headers["COP_X_mm"],
                  "3x Interpolated ~60Hz", ws_interp.max_row)
    ws_summary = wb["Summary"]
    ws_summary.add_chart(chart_x, "D2")

    # COP-Y vs time
    chart_y = ScatterChart()
    chart_y.title = "COP-Y: Raw ~20Hz vs 3x Interpolated ~60Hz"
    chart_y.x_axis.title = "Time (s)"
    chart_y.y_axis.title = "COP-Y (mm)"
    chart_y.height = 10
    chart_y.width = 18
    add_xy_series(chart_y, ws_raw, raw_headers["relative_time_s"], raw_headers["COP_Y_mm"],
                  "Raw ~20Hz", ws_raw.max_row, marker="circle")
    add_xy_series(chart_y, ws_interp, interp_headers["relative_time_s"], interp_headers["COP_Y_mm"],
                  "3x Interpolated ~60Hz", ws_interp.max_row)
    ws_summary.add_chart(chart_y, "D22")

    # COP trajectory
    chart_xy = ScatterChart()
    chart_xy.title = "COP Trajectory: Raw ~20Hz vs 3x Interpolated ~60Hz"
    chart_xy.x_axis.title = "COP-X (mm)"
    chart_xy.y_axis.title = "COP-Y (mm)"
    chart_xy.height = 12
    chart_xy.width = 18
    add_xy_series(chart_xy, ws_raw, raw_headers["COP_X_mm"], raw_headers["COP_Y_mm"],
                  "Raw ~20Hz", ws_raw.max_row, marker="circle")
    add_xy_series(chart_xy, ws_interp, interp_headers["COP_X_mm"], interp_headers["COP_Y_mm"],
                  "3x Interpolated ~60Hz", ws_interp.max_row)
    ws_summary.add_chart(chart_xy, "D42")


# ============================================================
# 顯示結果
# ============================================================

print()
print("========================================")
print("完成")
print("========================================")

print(
    f"Raw Samples       : "
    f"{len(raw_df)}"
)

print(
    f"Raw Duration      : "
    f"{raw_duration:.3f} s"
)

print(
    f"Raw Actual Hz     : "
    f"{raw_actual_hz:.2f} Hz"
)

print()

print(
    f"3x Samples        : "
    f"{len(interp_df)}"
)

print(
    f"Interpolated Hz   : "
    f"{interp_actual_hz:.2f} Hz"
)

print()

print(
    "Excel 已輸出："
    + OUTPUT_FILE
)